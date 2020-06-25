#!/usr/bin/python
import sys
import csv

from openpyxl import load_workbook

survey_mapping = {
    "construction": {
        "phone_number": "0300 1234 910",
        "survey_name": "Monthly Business Survey - Construction and Allied Trades"
    },
    "qifdi": {
        "phone_number": "0300 1234 931",
        "survey_name": "Quarterly Inward Foreign Direct Investment Survey"
    },
    "qofdi": {
        "phone_number": "0300 1234 931",
        "survey_name": "Quarterly Outward Foreign Direct Investment Survey"
    }
}


def process_file(params):
    wb = load_workbook(filename=params['input_file'])
    sheet = wb.active

    print_file_dict = create_print_file_dictionary(params['print_file'])

    with open('enrolment_email_details.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['email address', 'RU_NAME', 'SURVEY_NAME',
                         'RESPONDENT_PERIOD', 'RETURN_BY_DATE', 'ENROLMENT_CODE', 'SURVEY_PHONELINE'])

        for row in sheet.iter_rows(min_row=2):
            try:
                enrolment_code = print_file_dict[str(row[0].value)]
            except KeyError:
                print(f"ru_ref [{row[0].value}] is present in the input file but absent in the printfile.  Skipping.")
                continue

            ru_name = row[1].value
            email_address = row[2].value

            if email_address:
                if "@" in email_address:
                    csv_line = [email_address, ru_name, params['survey_name'], params['respondent_period'],
                                params['return_by_date'], enrolment_code, params['survey_phone_number']]
                    writer.writerow(csv_line)
                else:
                    print(f"Email address {email_address} isn't a valid email address. Skipping")
                    continue


def create_print_file_dictionary(print_file):
    """Takes the printfile, reads it and returns a dictionary where the keys are ru_refs and the values
    are the enrolment codes
    """
    print_file_dict = {}
    with open(params['print_file'], newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=':')
        for row in reader:
            print_file_dict[row[0]] = row[1]
    return print_file_dict


if __name__ == "__main__":
    if len(sys.argv) != 6:  # First argument is the name of the script
        length = len(sys.argv)
        sys.exit(f"Exactly 6 arguments expected. Number provided [{length}]")
 
    survey = sys.argv[3]
    if not survey_mapping[survey.lower()]:
        sys.exit(f"Survey [{survey}] not supported by this script")
        
    params = {
        'input_file': sys.argv[1],
        'print_file': sys.argv[2],
        'survey_name': survey_mapping[survey.lower()]['survey_name'],
        'respondent_period': sys.argv[4],
        'return_by_date': sys.argv[5],
        'survey_phone_number': survey_mapping[survey]['phone_number']
    }

    print(f"Input file is {params['input_file']}")
    print(f"Print file is {params['print_file']}")
    
    process_file(params)
